import discord
import os
import asyncio
import random
import openpyxl
import Image
import datetime

client = discord.Client()


@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("---------------------")
    await client.change_presence(game=discord.Game(name='실험체 도움말 치셈', type=1))


@client.event
async def on_member_join(member):
    role = ""
    for i in member.server.roles:
        if i.name == "gamer":
            role = i
            break
    await client.add_roles(member, role)


@client.event
async def on_message(message):
    if message.content.startswith('안녕 실험체'):
        file = openpyxl.load_workbook("쿨타임(안녕 실험체).xlsx")
        sheet = file.active
        for i in range(1, 101):
            if sheet["A" + str(i)].value == message.author.id:
                if int(sheet["B" + str(i)].value) <= int(datetime.datetime.today().strftime("%V%m%d%H%M%S")):
                    await client.send_message(message.channel, "ㅎㅇ?")
                    a = datetime.datetime.today() + datetime.timedelta(seconds=5)
                    sheet["B" + str(i)].value = a.strftime("%V%m%d%H%M%S")
                    file.save("쿨타임(안녕 실험체).xlsx")
                else:
                    await client.send_message(message.channel, "좀 천천히해 인사하기도 힘드네;")
                break
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = message.author.id
                a = datetime.datetime.today() + datetime.timedelta(seconds=5)
                sheet["B" + str(i)].value = a.strftime("%V%m%d%H%M%S")
                await client.send_message(message.channel, "ㅎㅇ?")
                file.save("쿨타임(안녕 실험체).xlsx")
                break

    if message.content.startswith('실험체 주사위굴려'):
        roll = message.content.split(" ")
        rolled = roll[1].split("d")
        dice = 0
        for i in range(1, int(rolled[0]) + 1):
            dice = dice + random.randint(1, int(rolled[1]))
        await client.send_message(message.channel, str(dice))

    if message.content.startswith('실험체 뭐가나아'):
        choice = message.content.split(" ")
        choicenumber = random.randint(1, len(choice) - 1)
        choiceresult = choice[choicenumber]
        await client.send_message(message.channel, choiceresult)

    if message.content.startswith('오늘의 추천메뉴'):
        food = "짜장면 짬뽕 라면 스테이크 침 치킨 족발 순대 양념치킨 포테이토칩 담배100갑"
        foodchoice = food.split(" ")
        foodnumber = random.randint(1, len(foodchoice))
        foodresult = foodchoice[foodnumber - 1]
        await client.send_message(message.channel, foodresult)

    if message.content.startswith('실험체 메모해'):
        file = open("실험체용메모장.txt", "w")
        file.write("안녕하세요")
        file.close()

    if message.content.startswith("실험체 메모읽어"):
        file = open("실험체용메모장.txt")
        await client.send_message(message.channel, file.read())
        file.close()

    if message.content.startswith("실험체 학습해"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                break
        file.save("기억.xlsx")

    if message.content.startswith("실험체 말해봐") and not message.content.startswith("잊어"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith("실험체 잊어"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await client.send_message(message.channel, "그거 까먹음")
                file.save("기억.xlsx")
                break

    if message.content.startswith('실험체 팀나눠'):
        team = message.content[8:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel, person[i] + " 은(는) " + teamname[i])

    if message.content.startswith("실험체 검색해"):
        img = message.content.split(" ")
        imgsrc = Image.get_image(img[1])
        await client.send_message(message.channel, imgsrc)

    if message.content.startswith("실험체 도움말"):
        await client.send_message(message.channel, "안녕 실험체 ")
        await client.send_message(message.channel, "실험체 주사위굴려 숫자d숫자")
        await client.send_message(message.channel, "실험체 뭐가나아 A B ")
        await client.send_message(message.channel, "오늘의 추천메뉴")
        await client.send_message(message.channel, "실험체 학습해 (~라고 플레이어가 말하면) (~라고 실험체가 대답)")
        await client.send_message(message.channel, "실험체 말해봐 (학습한것중 플레이어가 말하는 칸에 적는걸 적어야함)")
        await client.send_message(message.channel, "실험체 잊어")
        await client.send_message(message.channel, "실험체 팀나눠 A B C D E/1 2 3 4 5")


access_token = os.environ["Bot Token"]
client.run(access_token)
